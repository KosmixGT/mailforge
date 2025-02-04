from typing import List

from fastapi import APIRouter, HTTPException, status, UploadFile, File
from fastapi import Depends
from sqlalchemy.orm import Session
from app.database import get_db
from datetime import datetime

from app.mailing.schemas import MailingBase, MailingCreate
import app.mailing.crud as crud

from app.auth import get_current_user
from app.models import User

from io import BytesIO
from io import StringIO
import csv
import openpyxl

router = APIRouter(prefix='/mailings', tags=['mailing'])

@router.get("/{mailing_id}", response_model=MailingBase)
def get_mailing(mailing_id: int, db: Session = Depends(get_db)):
    mailing = crud.get_mailing(db, mailing_id)
    if not mailing:
        raise HTTPException(status_code=400, detail="Mailing does not exist")
    return mailing

@router.get("/by_user/{user_id}", response_model=List[MailingBase])
def get_mailings_by_user_id(user_id: int, db: Session = Depends(get_db)):
    return crud.get_mailings_by_user_id(db, user_id)

@router.get("/", response_model=List[MailingBase])  
def get_mailings(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return crud.get_mailings(db)

@router.post("/create", response_model=MailingCreate)
def create_mailing(mailing: MailingCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    return crud.create_mailing(db, mailing, current_user)

@router.put("/update/{mailing_id}", response_model=MailingCreate)
def update_mailing(mailing_id: int, mailing: MailingCreate, db: Session = Depends(get_db)):
    return crud.update_mailing(db, mailing_id, mailing)

@router.delete("/delete/{mailing_id}")
def delete_mailing(mailing_id: int, db: Session = Depends(get_db)):
    return crud.delete_mailing(db, mailing_id)

#эндпоинт для загрузки рассылок через csv/excel файл
@router.post("/upload")
async def upload_file(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    # Проверяем расширение файла
    if not file.filename.endswith(('.csv', '.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Допустимые форматы файлов: CSV, XLS, XLSX")
    
    contents = await file.read()  # Читаем содержимое файла
    data = parse_file_contents(contents, file.filename)
    for item in data:
        crud.create_mailing(db, item, current_user)
    return {"message": "Файл успешно загружен и данные добавлены в базу данных"}

def parse_file_contents(contents: bytes, filename: str) -> List[MailingCreate]:
    # Парсим содержимое файла в объекты MailingCreate
    data = []
    if filename.endswith('.csv'):
        data = parse_csv(contents)
    elif filename.endswith(('.xls', '.xlsx')):
        data = parse_excel(contents)
    return data

# Парсинг csv file
def parse_csv(contents: bytes) -> List[MailingCreate]:
    data = []

    contents_str = contents.decode('utf-8-sig')
    csv_file = StringIO(contents_str)

    reader = csv.reader(csv_file)
    next(reader)
    for row in reader:
        row = row[0].split('","')
        row[0] = row[0].lstrip('"')
        row[-1] = row[-1].rstrip('"')
        if len(row) >=2:
            # Извлекаем значения из строки
            title = row[0]
            text = row[1]
            scheduledtime_str = row[2]

            # Преобразуем строку времени в объект datetime, если есть значение
            scheduledtime = None
            if scheduledtime_str:
                scheduledtime = datetime.strptime(scheduledtime_str, '%Y-%m-%d %H:%M:%S')

            # Добавляем данные в список
            data.append(MailingCreate(title=title, messagetext=text, scheduledtime=scheduledtime))
        else:
            raise HTTPException(status_code=400, detail="Неверный формат данных в файле CSV")

    return data


# Парсинг excel file
def parse_excel(contents: bytes) -> List[MailingCreate]:
    data = []

    # Преобразуем содержимое файла в файл Excel с помощью openpyxl
    workbook = openpyxl.load_workbook(filename=BytesIO(contents))
    sheet = workbook.active

    # Пропускаем заголовок
    iter_rows = iter(sheet.rows)
    next(iter_rows)

    # Проходим по строкам в Excel файле и добавляем данные в список
    for row in iter_rows:
        title = row[0].value
        messagetext = row[1].value
        scheduledtime = row[2].value

        # Если scheduledtime представлен в виде строки, преобразуем его в объект datetime
        if isinstance(scheduledtime, str):
            scheduledtime = datetime.strptime(scheduledtime, '%Y-%m-%d %H:%M:%S')
        # Проверяем, что title и messagetext не равны None перед созданием объекта MailingCreate
        if title is not None and messagetext is not None:
            data.append(MailingCreate(title=title, messagetext=messagetext, scheduledtime=scheduledtime))
        else:
            # Обработка случая, когда title или messagetext равны None
            raise HTTPException(status_code=400, detail="Неверный формат данных в файле Excel")

    return data