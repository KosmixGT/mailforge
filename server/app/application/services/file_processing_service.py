from datetime import datetime
from typing import List
from io import StringIO, BytesIO
import csv
import openpyxl
from fastapi import HTTPException
from app.application.dto.mailing import MailingCreateDTO


class FileProcessingService:
    async def process_file(
        self, contents: bytes, filename: str
    ) -> List[MailingCreateDTO]:
        if filename.endswith(".csv"):
            return await self._parse_csv(contents)
        elif filename.endswith((".xls", ".xlsx")):
            return await self._parse_excel(contents)
        raise ValueError("Unsupported file format")

    async def _parse_csv(self, contents: bytes) -> List[MailingCreateDTO]:
        data = []
        contents_str = contents.decode("utf-8-sig")
        csv_file = StringIO(contents_str)

        reader = csv.reader(csv_file)
        next(reader)  # Skip header

        for row in reader:
            row = row[0].split('","')
            row[0] = row[0].lstrip('"')
            row[-1] = row[-1].rstrip('"')

            if len(row) >= 2:
                title = row[0]
                text = row[1]
                scheduledtime_str = row[2]

                scheduledtime = None
                if scheduledtime_str:
                    scheduledtime = datetime.strptime(
                        scheduledtime_str, "%Y-%m-%d %H:%M:%S"
                    )

                data.append(
                    MailingCreateDTO(
                        title=title, message_text=text, scheduled_time=scheduledtime
                    )
                )
            else:
                raise HTTPException(
                    status_code=400, detail="Неверный формат данных в файле CSV"
                )

        return data

    async def _parse_excel(self, contents: bytes) -> List[MailingCreateDTO]:
        data = []
        workbook = openpyxl.load_workbook(filename=BytesIO(contents))
        sheet = workbook.active

        iter_rows = iter(sheet.rows)
        next(iter_rows)  # Skip header

        for row in iter_rows:
            title = row[0].value
            messagetext = row[1].value
            scheduledtime = row[2].value

            if isinstance(scheduledtime, str):
                scheduledtime = datetime.strptime(scheduledtime, "%Y-%m-%d %H:%M:%S")

            if title is not None and messagetext is not None:
                data.append(
                    MailingCreateDTO(
                        title=title,
                        message_text=messagetext,
                        scheduled_time=scheduledtime,
                    )
                )
            else:
                raise HTTPException(
                    status_code=400, detail="Неверный формат данных в файле Excel"
                )

        return data
